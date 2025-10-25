import openpyxl
from openpyxl import Workbook

class ExcelHandler:
    def export_template(self, file_path):
        '''テンプレートExcelファイルを作成'''
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "バーコードリスト"
            
            # ヘッダー
            ws['A1'] = 'バーコード値'
            ws['B1'] = '表示名'
            
            # サンプルデータ
            ws['A2'] = '123456789'
            ws['B2'] = 'サンプル商品A'
            ws['A3'] = '987654321'
            ws['B3'] = 'サンプル商品B'
            
            # 列幅を調整
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"テンプレート作成エラー: {e}")
            return False
    
    def import_excel(self, file_path):
        '''Excelファイルからデータを読み込む'''
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # 両方の列にデータがある場合のみ
                    data.append({
                        'barcode': str(row[0]),
                        'name': str(row[1])
                    })
            
            return data
        except Exception as e:
            print(f"Excel読み込みエラー: {e}")
            return []
