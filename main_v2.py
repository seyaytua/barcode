import sys
import os
from datetime import datetime
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
                               QFileDialog, QMessageBox, QTabWidget, QLabel,
                               QDateEdit, QGroupBox, QComboBox, QSpinBox, QDialog,
                               QScrollArea)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QPixmap, QImage
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from barcode import Code128
    from barcode.writer import ImageWriter
except ImportError:
    print("pip3 install python-barcode pillow")
    sys.exit(1)

try:
    import qrcode
except ImportError:
    print("pip3 install qrcode[pil]")
    sys.exit(1)

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
import tempfile


class PreviewDialog(QDialog):
    def __init__(self, preview_images, parent=None):
        super().__init__(parent)
        self.setWindowTitle("プレビュー")
        self.setGeometry(100, 100, 900, 700)
        self.preview_images = preview_images
        self.current_page = 0
        
        layout = QVBoxLayout(self)
        
        nav = QHBoxLayout()
        self.prev_btn = QPushButton("◀ 前")
        self.prev_btn.clicked.connect(self.prev_page)
        self.page_label = QLabel(f"1 / {len(preview_images)}")
        self.page_label.setAlignment(Qt.AlignCenter)
        self.next_btn = QPushButton("次 ▶")
        self.next_btn.clicked.connect(self.next_page)
        nav.addWidget(self.prev_btn)
        nav.addWidget(self.page_label)
        nav.addWidget(self.next_btn)
        layout.addLayout(nav)
        
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.img_label = QLabel()
        self.img_label.setAlignment(Qt.AlignCenter)
        self.scroll.setWidget(self.img_label)
        layout.addWidget(self.scroll)
        
        close_btn = QPushButton("閉じる")
        close_btn.clicked.connect(self.close)
        close_btn.setMinimumHeight(40)
        layout.addWidget(close_btn)
        
        self.update_display()
    
    def update_display(self):
        if self.preview_images:
            img = self.preview_images[self.current_page]
            qimg = QImage(img.tobytes(), img.width, img.height, img.width * 3, QImage.Format_RGB888)
            self.img_label.setPixmap(QPixmap.fromImage(qimg))
            self.page_label.setText(f"{self.current_page + 1} / {len(self.preview_images)}")
        self.prev_btn.setEnabled(self.current_page > 0)
        self.next_btn.setEnabled(self.current_page < len(self.preview_images) - 1)
    
    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_display()
    
    def next_page(self):
        if self.current_page < len(self.preview_images) - 1:
            self.current_page += 1
            self.update_display()


class BarcodeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("バーコード管理システム")
        self.setGeometry(100, 100, 1200, 800)
        
        self.equipment_data = []
        self.teacher_data = []
        self.creation_date = datetime.now()
        self.code_type = "barcode"
        self.font_size = 10
        
        self.center_margin = 4
        self.page_margin = 8
        
        self._font_path = None
        self._init_fonts()
        
        try:
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        except:
            pass
        
        self.init_ui()
    
    def _init_fonts(self):
        paths = [
            "/System/Library/Fonts/Helvetica.ttc",
            "/System/Library/Fonts/HelveticaNeue.ttc",
            "/Library/Fonts/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for p in paths:
            if os.path.exists(p):
                try:
                    ImageFont.truetype(p, 12)
                    self._font_path = p
                    return
                except:
                    pass
    
    def _get_font(self, size):
        size = max(8, int(size))
        if self._font_path:
            try:
                return ImageFont.truetype(self._font_path, size)
            except:
                pass
        return ImageFont.load_default()
    
    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        
        layout.addLayout(self.create_header())
        
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        
        self.tabs.addTab(self.create_equipment_tab(), "備品用バーコード")
        self.tabs.addTab(self.create_teacher_tab(), "教員用バーコード")
        
        layout.addLayout(self.create_buttons())
    
    def create_header(self):
        layout = QVBoxLayout()
        
        title = QLabel("バーコード/QRコード台紙作成システム")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        group = QGroupBox("設定")
        h = QHBoxLayout()
        
        h.addWidget(QLabel("作成日付:"))
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.dateChanged.connect(self.update_date)
        h.addWidget(self.date_edit)
        
        h.addSpacing(20)
        h.addWidget(QLabel("コードタイプ:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["バーコード", "QRコード"])
        self.type_combo.currentIndexChanged.connect(self.update_type)
        h.addWidget(self.type_combo)
        
        h.addSpacing(20)
        h.addWidget(QLabel("番号サイズ:"))
        self.font_spin = QSpinBox()
        self.font_spin.setRange(8, 20)
        self.font_spin.setValue(10)
        self.font_spin.setSuffix(" pt")
        self.font_spin.valueChanged.connect(lambda v: setattr(self, 'font_size', v))
        h.addWidget(self.font_spin)
        h.addStretch()
        
        group.setLayout(h)
        layout.addWidget(group)
        return layout
    
    def create_equipment_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.eq_label = QLabel("備品用（サイズ1-2:8列, 3-4:6列, 5-6:4列, 7-8:2列）")
        layout.addWidget(self.eq_label)
        self.eq_table = QTableWidget()
        self.eq_table.setColumnCount(4)
        self.eq_table.setHorizontalHeaderLabels(["対象名", "コード番号", "サイズ", "備考"])
        self.eq_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.eq_table)
        return tab
    
    def create_teacher_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.tc_label = QLabel("教員用（2列レイアウト）")
        layout.addWidget(self.tc_label)
        self.tc_table = QTableWidget()
        self.tc_table.setColumnCount(3)
        self.tc_table.setHorizontalHeaderLabels(["教員名", "コード番号", "備考"])
        self.tc_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.tc_table)
        return tab
    
    def create_buttons(self):
        main = QVBoxLayout()
        
        row1 = QHBoxLayout()
        btn = QPushButton("📥 テンプレート出力")
        btn.clicked.connect(self.export_template)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#FF9800;color:white;")
        row1.addWidget(btn)
        btn = QPushButton("📂 Excel読み込み")
        btn.clicked.connect(self.load_excel)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#4CAF50;color:white;")
        row1.addWidget(btn)
        main.addLayout(row1)
        
        row2 = QHBoxLayout()
        btn = QPushButton("👁️ 備品プレビュー")
        btn.clicked.connect(self.preview_equipment)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#9C27B0;color:white;")
        row2.addWidget(btn)
        btn = QPushButton("👁️ 教員プレビュー")
        btn.clicked.connect(self.preview_teacher)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#9C27B0;color:white;")
        row2.addWidget(btn)
        main.addLayout(row2)
        
        row3 = QHBoxLayout()
        btn = QPushButton("🖨️ 備品PDF出力")
        btn.clicked.connect(self.print_equipment)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#2196F3;color:white;")
        row3.addWidget(btn)
        btn = QPushButton("🖨️ 教員PDF出力")
        btn.clicked.connect(self.print_teacher)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#2196F3;color:white;")
        row3.addWidget(btn)
        btn = QPushButton("🗑️ クリア")
        btn.clicked.connect(self.clear_data)
        btn.setMinimumHeight(45)
        btn.setStyleSheet("font-size:14px;font-weight:bold;background:#f44336;color:white;")
        row3.addWidget(btn)
        main.addLayout(row3)
        
        return main
    
    def update_date(self, d):
        self.creation_date = datetime(d.year(), d.month(), d.day())
    
    def update_type(self, i):
        self.code_type = "barcode" if i == 0 else "qrcode"
    
    def get_columns(self, size):
        """サイズに応じた全体の列数"""
        if size <= 2:
            return 8  # 左4+右4
        elif size <= 4:
            return 6  # 左3+右3
        elif size <= 6:
            return 4  # 左2+右2
        else:
            return 2  # 左1+右1
    
    def get_cell_height(self, size):
        if self.code_type == "barcode":
            return 35 + size * 8
        else:
            return 40 + size * 10
    
    def crosses_center(self, start, end, center, margin):
        return not (end <= center - margin or start >= center + margin)
    
    def calc_positions(self, pw, ph, margin, header_h, cell_h, columns):
        """中心線を避けてセル位置を計算"""
        cx, cy = pw / 2, ph / 2
        cm = self.center_margin
        
        left_cols = columns // 2
        right_cols = columns - left_cols
        
        left_w = cx - margin - cm
        right_w = pw - cx - cm - margin
        
        left_cell_w = left_w / left_cols if left_cols > 0 else 0
        right_cell_w = right_w / right_cols if right_cols > 0 else 0
        
        positions = []
        y = ph - margin - header_h
        
        while y - cell_h >= margin:
            top, bottom = y, y - cell_h
            
            if self.crosses_center(bottom, top, cy, cm):
                y = cy - cm
                continue
            
            # 左側
            for i in range(left_cols):
                positions.append({
                    'x': margin + i * left_cell_w,
                    'y': top,
                    'w': left_cell_w,
                    'h': cell_h
                })
            
            # 右側
            for i in range(right_cols):
                positions.append({
                    'x': cx + cm + i * right_cell_w,
                    'y': top,
                    'w': right_cell_w,
                    'h': cell_h
                })
            
            y -= cell_h
        
        return positions
    
    def export_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存", "テンプレート.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            hf = Font(bold=True, size=11, color="FFFFFF")
            hfill = PatternFill("solid", fgColor="4472C4")
            
            # 備品用
            ws = wb.create_sheet("備品用", 0)
            for i, h in enumerate(["対象名", "コード番号", "サイズ", "備考"], 1):
                c = ws.cell(1, i, h)
                c.font = hf
                c.fill = hfill
            for r, d in enumerate([["PC-001", "10000001", 3, ""], ["プリンタ", "10000002", 4, ""]], 2):
                for c, v in enumerate(d, 1):
                    ws.cell(r, c, v)
            
            # 教員用
            ws = wb.create_sheet("教員用", 1)
            for i, h in enumerate(["教員名", "コード番号", "備考"], 1):
                c = ws.cell(1, i, h)
                c.font = hf
                c.fill = hfill
            for r, d in enumerate([["山田太郎", "20000001", ""], ["佐藤花子", "20000002", ""]], 2):
                for c, v in enumerate(d, 1):
                    ws.cell(r, c, v)
            
            wb.save(path)
            QMessageBox.information(self, "成功", f"作成完了:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "エラー", str(e))
    
    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "開く", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            loaded = []
            
            if "備品用" in wb.sheetnames:
                self.equipment_data = []
                self.eq_table.setRowCount(0)
                for row in wb["備品用"].iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    name = str(row[0])
                    code = str(row[1]) if row[1] else ""
                    try:
                        size = int(row[2]) if row[2] else 3
                    except:
                        size = 3
                    size = max(1, min(8, size))
                    note = str(row[3]) if len(row) > 3 and row[3] else ""
                    
                    self.equipment_data.append({'name': name, 'code': code, 'size': size, 'note': note})
                    r = self.eq_table.rowCount()
                    self.eq_table.insertRow(r)
                    self.eq_table.setItem(r, 0, QTableWidgetItem(name))
                    self.eq_table.setItem(r, 1, QTableWidgetItem(code))
                    self.eq_table.setItem(r, 2, QTableWidgetItem(str(size)))
                    self.eq_table.setItem(r, 3, QTableWidgetItem(note))
                loaded.append("備品用")
            
            if "教員用" in wb.sheetnames:
                self.teacher_data = []
                self.tc_table.setRowCount(0)
                for row in wb["教員用"].iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    name = str(row[0])
                    code = str(row[1]) if row[1] else ""
                    note = str(row[2]) if len(row) > 2 and row[2] else ""
                    
                    self.teacher_data.append({'name': name, 'code': code, 'note': note})
                    r = self.tc_table.rowCount()
                    self.tc_table.insertRow(r)
                    self.tc_table.setItem(r, 0, QTableWidgetItem(name))
                    self.tc_table.setItem(r, 1, QTableWidgetItem(code))
                    self.tc_table.setItem(r, 2, QTableWidgetItem(note))
                loaded.append("教員用")
            
            if loaded:
                QMessageBox.information(self, "成功", f"読み込み: {', '.join(loaded)}")
            else:
                QMessageBox.warning(self, "警告", "シートが見つかりません")
        except Exception as e:
            QMessageBox.critical(self, "エラー", str(e))
    
    def gen_barcode(self, code, scale=1.0):
        try:
            bc = Code128(str(code), writer=ImageWriter())
            buf = BytesIO()
            bc.write(buf, {
                'module_width': 0.3 * scale,
                'module_height': 12 * scale,
                'font_size': 0,
                'text_distance': 1,
                'quiet_zone': 2,
                'write_text': False,
            })
            buf.seek(0)
            img = Image.open(buf).convert('RGB')
            
            # テキスト追加
            text = str(code)
            font = self._get_font(max(8, int(self.font_size * scale)))
            
            w, h = img.size
            try:
                tmp = ImageDraw.Draw(Image.new('RGB', (1, 1)))
                bb = tmp.textbbox((0, 0), text, font=font)
                tw, th = bb[2] - bb[0], bb[3] - bb[1]
            except:
                tw, th = len(text) * 6, 10
            
            new = Image.new('RGB', (w, h + th + 3), 'white')
            new.paste(img, (0, 0))
            d = ImageDraw.Draw(new)
            try:
                d.text(((w - tw) / 2, h + 1), text, fill='black', font=font)
            except:
                d.text(((w - tw) / 2, h + 1), text, fill='black')
            return new
        except Exception as e:
            print(f"BC error: {e}")
            return None
    
    def gen_qr(self, code, scale=1.0):
        try:
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L,
                               box_size=max(2, int(5 * scale)), border=1)
            qr.add_data(str(code))
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            return img.convert('RGB') if hasattr(img, 'convert') else img
        except Exception as e:
            print(f"QR error: {e}")
            return None
    
    def create_eq_pages(self):
        if not self.equipment_data:
            return []
        
        pages = []
        pw, ph = 595, 842
        margin = self.page_margin
        hdr_h = 28
        cx, cy = pw / 2, ph / 2
        
        title_font = self._get_font(9)
        text_font = self._get_font(7)
        
        idx = 0
        pnum = 0
        
        while idx < len(self.equipment_data):
            pnum += 1
            img = Image.new('RGB', (pw, ph), 'white')
            draw = ImageDraw.Draw(img)
            
            # ヘッダー
            ct = "BC" if self.code_type == "barcode" else "QR"
            try:
                draw.text((margin, 3), f"備品用{ct}台紙 {self.creation_date.strftime('%Y/%m/%d')} P.{pnum}", fill='black', font=title_font)
            except:
                draw.text((margin, 3), f"P.{pnum}", fill='black')
            draw.line((margin, 18, pw - margin, 18), fill='black')
            draw.line((int(cx), 20, int(cx), ph - margin), fill='#EEEEEE')
            draw.line((margin, int(cy), pw - margin, int(cy)), fill='#EEEEEE')
            
            cur_size = self.equipment_data[idx]['size']
            cols = self.get_columns(cur_size)
            cell_h = self.get_cell_height(cur_size)
            positions = self.calc_positions(pw, ph, margin, hdr_h, cell_h, cols)
            
            pi = 0
            while idx < len(self.equipment_data) and pi < len(positions):
                item = self.equipment_data[idx]
                
                if item['size'] != cur_size:
                    cur_size = item['size']
                    cols = self.get_columns(cur_size)
                    cell_h = self.get_cell_height(cur_size)
                    
                    rem_y = positions[pi - 1]['y'] - positions[pi - 1]['h'] if pi > 0 else ph - margin - hdr_h
                    positions = self.calc_positions(pw, ph, margin, hdr_h, cell_h, cols)
                    positions = [p for p in positions if p['y'] <= rem_y]
                    pi = 0
                    if not positions:
                        break
                
                p = positions[pi]
                x, y = p['x'], ph - p['y']
                w, h = p['w'], p['h']
                
                draw.rectangle([x, y, x + w, y + h], outline='#AAAAAA')
                
                name = item['name'][:int(w / 7)]
                try:
                    draw.text((x + 2, y + 1), name, fill='black', font=text_font)
                except:
                    draw.text((x + 2, y + 1), name, fill='black')
                
                scale = 0.5 + item['size'] * 0.12
                if self.code_type == "barcode":
                    code_img = self.gen_barcode(item['code'], scale)
                    if code_img:
                        cw = min(int(w - 4), code_img.width)
                        ch = min(int(h - 12), code_img.height)
                        code_img = code_img.resize((cw, ch), Image.Resampling.LANCZOS)
                        img.paste(code_img, (int(x + (w - cw) / 2), int(y + 10)))
                else:
                    code_img = self.gen_qr(item['code'], scale)
                    if code_img:
                        sz = min(int(w - 4), int(h - 12))
                        code_img = code_img.resize((sz, sz), Image.Resampling.LANCZOS)
                        img.paste(code_img, (int(x + (w - sz) / 2), int(y + 10)))
                
                idx += 1
                pi += 1
            
            pages.append(img)
        
        return pages
    
    def preview_equipment(self):
        if not self.equipment_data:
            QMessageBox.warning(self, "警告", "データがありません")
            return
        try:
            pages = self.create_eq_pages()
            if pages:
                PreviewDialog(pages, self).exec()
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "エラー", traceback.format_exc())
    
    def create_tc_pages(self):
        if not self.teacher_data:
            return []
        
        pages = []
        pw, ph = 595, 842
        margin = self.page_margin
        hdr_h = 28
        cx, cy = pw / 2, ph / 2
        cm = self.center_margin
        
        title_font = self._get_font(9)
        text_font = self._get_font(8)
        
        row_h = 32 if self.code_type == "barcode" else 38
        col_w = cx - margin - cm
        
        positions = []
        y = ph - margin - hdr_h
        while y - row_h >= margin:
            if self.crosses_center(y - row_h, y, cy, cm):
                y = cy - cm
                continue
            positions.append({'x': margin, 'y': y})
            positions.append({'x': cx + cm, 'y': y})
            y -= row_h
        
        per_page = len(positions)
        
        for start in range(0, len(self.teacher_data), per_page):
            pnum = start // per_page + 1
            img = Image.new('RGB', (pw, ph), 'white')
            draw = ImageDraw.Draw(img)
            
            ct = "BC" if self.code_type == "barcode" else "QR"
            try:
                draw.text((margin, 3), f"教員用{ct}台紙 {self.creation_date.strftime('%Y/%m/%d')} P.{pnum}", fill='black', font=title_font)
            except:
                draw.text((margin, 3), f"P.{pnum}", fill='black')
            draw.line((margin, 18, pw - margin, 18), fill='black')
            draw.line((int(cx), 20, int(cx), ph - margin), fill='#EEEEEE')
            draw.line((margin, int(cy), pw - margin, int(cy)), fill='#EEEEEE')
            
            items = self.teacher_data[start:start + per_page]
            for i, item in enumerate(items):
                if i >= len(positions):
                    break
                p = positions[i]
                x, y = p['x'], ph - p['y']
                
                try:
                    draw.text((x + 2, y + 3), item['name'][:8], fill='black', font=text_font)
                except:
                    draw.text((x + 2, y + 3), item['name'][:8], fill='black')
                
                if self.code_type == "barcode":
                    code_img = self.gen_barcode(item['code'], 0.5)
                    if code_img:
                        code_img = code_img.resize((80, 24), Image.Resampling.LANCZOS)
                        img.paste(code_img, (int(x + 62), int(y + 2)))
                else:
                    code_img = self.gen_qr(item['code'], 0.45)
                    if code_img:
                        code_img = code_img.resize((28, 28), Image.Resampling.LANCZOS)
                        img.paste(code_img, (int(x + 75), int(y + 2)))
            
            pages.append(img)
        
        return pages
    
    def preview_teacher(self):
        if not self.teacher_data:
            QMessageBox.warning(self, "警告", "データがありません")
            return
        try:
            pages = self.create_tc_pages()
            if pages:
                PreviewDialog(pages, self).exec()
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "エラー", traceback.format_exc())
    
    def print_equipment(self):
        if not self.equipment_data:
            QMessageBox.warning(self, "警告", "データがありません")
            return
        name = "バーコード" if self.code_type == "barcode" else "QRコード"
        path, _ = QFileDialog.getSaveFileName(self, "保存", f"備品用{name}.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            self.create_eq_pdf(path)
            QMessageBox.information(self, "成功", f"作成完了:\n{path}")
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "エラー", traceback.format_exc())
    
    def create_eq_pdf(self, path):
        c = canvas.Canvas(path, pagesize=A4)
        w, h = A4
        margin = self.page_margin
        hdr_h = 28
        cx, cy = w / 2, h / 2
        
        ct = "BC" if self.code_type == "barcode" else "QR"
        
        def header(pn):
            try:
                c.setFont("HeiseiMin-W3", 8)
            except:
                c.setFont("Helvetica", 8)
            c.drawString(margin, h - 12, f"備品用{ct}台紙 {self.creation_date.strftime('%Y/%m/%d')} P.{pn}")
            c.line(margin, h - 18, w - margin, h - 18)
            c.setStrokeColorRGB(0.93, 0.93, 0.93)
            c.line(cx, margin, cx, h - hdr_h)
            c.line(margin, cy, w - margin, cy)
            c.setStrokeColorRGB(0, 0, 0)
        
        pn = 1
        header(pn)
        
        idx = 0
        while idx < len(self.equipment_data):
            cur_size = self.equipment_data[idx]['size']
            cols = self.get_columns(cur_size)
            cell_h = self.get_cell_height(cur_size)
            positions = self.calc_positions(w, h, margin, hdr_h, cell_h, cols)
            
            pi = 0
            while idx < len(self.equipment_data) and pi < len(positions):
                item = self.equipment_data[idx]
                
                if item['size'] != cur_size:
                    cur_size = item['size']
                    cols = self.get_columns(cur_size)
                    cell_h = self.get_cell_height(cur_size)
                    
                    rem_y = positions[pi - 1]['y'] - positions[pi - 1]['h'] if pi > 0 else h - margin - hdr_h
                    positions = self.calc_positions(w, h, margin, hdr_h, cell_h, cols)
                    positions = [p for p in positions if p['y'] <= rem_y]
                    pi = 0
                    if not positions:
                        break
                
                p = positions[pi]
                px, py = p['x'], p['y']
                pw, ph = p['w'], p['h']
                
                c.setStrokeColorRGB(0.67, 0.67, 0.67)
                c.rect(px, py - ph, pw, ph)
                c.setStrokeColorRGB(0, 0, 0)
                
                try:
                    c.setFont("HeiseiMin-W3", 6)
                except:
                    c.setFont("Helvetica", 6)
                c.drawString(px + 2, py - 8, item['name'][:int(pw / 6)])
                
                scale = 0.5 + item['size'] * 0.12
                if self.code_type == "barcode":
                    code_img = self.gen_barcode(item['code'], scale)
                    if code_img:
                        iw = min(pw - 4, code_img.width)
                        ih = min(ph - 10, code_img.height)
                        ix = px + (pw - iw) / 2
                        iy = py - ph + 2
                else:
                    code_img = self.gen_qr(item['code'], scale)
                    if code_img:
                        sz = min(pw - 4, ph - 10)
                        iw = ih = sz
                        ix = px + (pw - sz) / 2
                        iy = py - ph + 2
                
                if code_img:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as f:
                        code_img.save(f.name)
                        try:
                            c.drawImage(f.name, ix, iy, width=iw, height=ih, mask='auto')
                        finally:
                            os.unlink(f.name)
                
                idx += 1
                pi += 1
            
            if idx < len(self.equipment_data):
                c.showPage()
                pn += 1
                header(pn)
        
        c.save()
    
    def print_teacher(self):
        if not self.teacher_data:
            QMessageBox.warning(self, "警告", "データがありません")
            return
        name = "バーコード" if self.code_type == "barcode" else "QRコード"
        path, _ = QFileDialog.getSaveFileName(self, "保存", f"教員用{name}.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            self.create_tc_pdf(path)
            QMessageBox.information(self, "成功", f"作成完了:\n{path}")
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "エラー", traceback.format_exc())
    
    def create_tc_pdf(self, path):
        c = canvas.Canvas(path, pagesize=A4)
        w, h = A4
        margin = self.page_margin
        hdr_h = 28
        cx, cy = w / 2, h / 2
        cm = self.center_margin
        
        ct = "BC" if self.code_type == "barcode" else "QR"
        row_h = 30 if self.code_type == "barcode" else 35
        col_w = cx - margin - cm
        
        positions = []
        y = h - margin - hdr_h
        while y - row_h >= margin:
            if self.crosses_center(y - row_h, y, cy, cm):
                y = cy - cm
                continue
            positions.append({'x': margin, 'y': y})
            positions.append({'x': cx + cm, 'y': y})
            y -= row_h
        
        per_page = len(positions)
        pn = 0
        
        def header(pn):
            try:
                c.setFont("HeiseiMin-W3", 8)
            except:
                c.setFont("Helvetica", 8)
            c.drawString(margin, h - 12, f"教員用{ct}台紙 {self.creation_date.strftime('%Y/%m/%d')} P.{pn}")
            c.line(margin, h - 18, w - margin, h - 18)
            c.setStrokeColorRGB(0.93, 0.93, 0.93)
            c.line(cx, margin, cx, h - hdr_h)
            c.line(margin, cy, w - margin, cy)
            c.setStrokeColorRGB(0, 0, 0)
        
        for start in range(0, len(self.teacher_data), per_page):
            if start > 0:
                c.showPage()
            pn += 1
            header(pn)
            
            items = self.teacher_data[start:start + per_page]
            for i, item in enumerate(items):
                if i >= len(positions):
                    break
                p = positions[i]
                px, py = p['x'], p['y']
                
                try:
                    c.setFont("HeiseiMin-W3", 7)
                except:
                    c.setFont("Helvetica", 7)
                c.drawString(px + 2, py - 10, item['name'][:10])
                
                if self.code_type == "barcode":
                    code_img = self.gen_barcode(item['code'], 0.52)
                    iw, ih = 85, 24
                    ix = px + 62
                else:
                    code_img = self.gen_qr(item['code'], 0.48)
                    iw = ih = 26
                    ix = px + 78
                
                if code_img:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as f:
                        code_img.save(f.name)
                        try:
                            c.drawImage(f.name, ix, py - row_h + 3, width=iw, height=ih, mask='auto')
                        finally:
                            os.unlink(f.name)
        
        c.save()
    
    def clear_data(self):
        if QMessageBox.question(self, "確認", "クリアしますか?") == QMessageBox.Yes:
            self.equipment_data = []
            self.teacher_data = []
            self.eq_table.setRowCount(0)
            self.tc_table.setRowCount(0)


def main():
    app = QApplication(sys.argv)
    BarcodeApp().show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
